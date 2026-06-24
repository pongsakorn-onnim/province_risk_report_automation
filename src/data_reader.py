import openpyxl
import openpyxl.reader.excel as _excel_mod
import openpyxl.reader.drawings as _drawings_mod
from pathlib import Path

# Some Excel files have broken drawing references (missing 'xl/drawings/drawing1.xml').
# openpyxl.reader.excel imports find_images by name, so patch both the source module
# and excel.py's local reference.
_orig_find_images = _drawings_mod.find_images


def _safe_find_images(archive, path):
    try:
        return _orig_find_images(archive, path)
    except KeyError:
        return [], []


_drawings_mod.find_images = _safe_find_images
_excel_mod.find_images = _safe_find_images

# Province Excel TYPE values
TYPE_FLOOD = "เสี่ยงท่วม"
TYPE_DROUGHT = "เสี่ยงแล้ง"
TYPE_BOTH = "เสี่ยงทั้งท่วมทั้งแล้ง"

# National Excel TYPE_RISK_E values
NAT_FLOOD = "Flood"
NAT_DROUGHT = "Drought"
NAT_BOTH = "flood/drought"


def read_province_6m(excel_path: Path) -> dict:
    """
    Returns {TYPE: {amphoe, tambon, detail}} from summary_6m sheet.
    TYPE keys: เสี่ยงท่วม, เสี่ยงแล้ง, เสี่ยงทั้งท่วมทั้งแล้ง
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["summary_6m"]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # REGION_CODE, REGION, PROV_T, TYPE, province_count, amphoe_count, tambon_count, PROV_TH
        if len(row) < 8:
            continue
        risk_type = row[3]
        if risk_type in (TYPE_FLOOD, TYPE_DROUGHT, TYPE_BOTH):
            result[risk_type] = {
                "amphoe": row[5] or 0,
                "tambon": row[6] or 0,
                "detail": row[7] or "",
            }
    return result


def read_province_monthly(excel_path: Path) -> dict:
    """
    Returns {month_year_key: {TYPE: {amphoe, tambon, detail}}} from summary_1m sheet.
    month_year_key format: "MM-YYYY" (e.g. "06-2026")
    Only includes เสี่ยงท่วม and เสี่ยงแล้ง rows.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["summary_1m"]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # month_year, TYPE_RISK_T, province, amphoe, tambon, detail_text
        if len(row) < 6:
            continue
        month_year = row[0]  # e.g. "06-2026"
        risk_type = row[1]
        if risk_type not in (TYPE_FLOOD, TYPE_DROUGHT):
            continue
        if month_year not in result:
            result[month_year] = {}
        result[month_year][risk_type] = {
            "amphoe": row[3] or 0,
            "tambon": row[4] or 0,
            "detail": row[5] or "",
        }
    return result


def read_national_6m(excel_path: Path) -> dict:
    """
    Returns {flood/drought/both: {province, amphoe, tambon, detail}} from Summary_6m sheet.
    Aggregates across all regions nationally. Column order varies by period — resolved by header name.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Summary_6m"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    def _get(row, *names):
        for name in names:
            if name in col and col[name] < len(row):
                return row[col[name]]
        return None

    totals = {k: {"province": 0, "amphoe": 0, "tambon": 0, "details": []}
              for k in (NAT_FLOOD, NAT_DROUGHT, NAT_BOTH)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        risk_type = _get(row, "TYPE_RISK_E")
        if risk_type not in totals:
            continue
        totals[risk_type]["province"] += _get(row, "province_count") or 0
        totals[risk_type]["amphoe"]   += _get(row, "amphoe_count") or 0
        totals[risk_type]["tambon"]   += _get(row, "tambon_count") or 0
        detail = _get(row, "detial_province", "detail_province")
        region = _get(row, "REGION_TH", "REGION")
        if detail:
            totals[risk_type]["details"].append({
                "region":    str(region).strip() if region else "",
                "provinces": str(detail).strip(),
            })

    def _build(key):
        t = totals[key]
        return {
            "province": t["province"],
            "amphoe":   t["amphoe"],
            "tambon":   t["tambon"],
            "detail":   t["details"],  # list of {"region": ..., "provinces": ...}
        }

    return {
        "flood": _build(NAT_FLOOD),
        "drought": _build(NAT_DROUGHT),
        "both": _build(NAT_BOTH),
    }
